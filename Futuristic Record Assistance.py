
from tkinter import *    
import cv2
import os
import PIL.Image, PIL.ImageTk
import pyttsx3
import datetime
import speech_recognition as sr
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
#import ImageTools

from PIL import Image

engine = pyttsx3.init('sapi5')
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)

window = Tk()

global var
global var1

var = StringVar()
var1 = StringVar()


w = load_workbook(filename="file2.xlsx")
s = w.active
wb = Workbook()
ws = wb.active

def speak(audio):
    engine.say(audio)
    engine.runAndWait()

def wishme():
    hour = int(datetime.datetime.now().hour)
    if hour>=0 and hour<12:
        speak("Good Morning!")
    elif hour>=12 and hour<18:
        speak("Good Afternoon")
    else:
        speak("Good Evening")
    
    speak("I am shubham. Please tell me how may I help you ") 

def takeCommand():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        var.set("Listening...")
        window.update()
        print("Listening...")
        r.pause_threshold = 1
        r.energy_threshold = 400
        audio = r.listen(source)
    try:
        var.set("Recognizing...")
        window.update()
        print("Recognizing...")
        query = r.recognize_google(audio, language='en-in')
        print("User said:",query) 
    except Exception as e:
        print(e)
        print("Say that again please...")
        return "None"
        var1.set(query)
        window.update()
    return query

def play():
    btn2['state'] = 'disabled'
    btn1.configure(bg = 'orange')
    wishme()
    ws.append(["name","standard","rollno"])
    while True:
        btn1.configure(bg = 'orange')
        query = takeCommand().lower()
        
        if 'hello' in query:
            var.set('Hello Sir')
            window.update()
            speak("Hello Sir")
			
        elif 'thank you' in query:
            var.set("Welcome Sir")
            window.update()
            speak("Welcome Sir")

        elif 'enter student details' in query:
            #while True:
                #sr_no = sr_no + 1
            
            var.set('Name of the student')
            window.update()
            speak('Name of the student')
            name = takeCommand()
            
            var.set('standard in which he/she study')
            window.update()
            speak('standard in which he/she study')
            standard = takeCommand()
            
            var.set('Roll Number')
            window.update()
            speak('Roll number')
            rollno = takeCommand()
            
            var.set('Please stand still for a photo')
            window.update()
            speak('Please stand still for a photo')
            

            w= load_workbook(filename="file2.xlsx")
            s=w.active
            l=Image(r"C:\Users\SHUBHAM BANSAL\Desktop\img1.jpg")

            l.height=150
            l.width=200
            i=1
            while True:
                i=i+1
                s.add_image(l,"E"+str(i))
                break
             
            ws.append([name,standard,rollno,s])
            wb.save("file2.xlsx")
            
            var.set('Details are saved')
            window.update()
            speak('Details are saved')
            
        #elif 'show me details' in query:
            #var.set('Name: '+name+' Standard: '+ standard+' Roll No.: '+ rollno)
            #window.update()
            #wb.save("file2.xlsx")
            
        elif 'exit' in query:
            var.set("Bye sir")
            btn1.configure(bg = '#5C85FB')
            btn2['state'] = 'normal'
            window.update()
            speak("Bye sir")
            break    
def update(ind):
    frame = frames[(ind)%100]
    ind += 1
    label.configure(image=frame)
    window.after(100, update, ind)

label2 = Label(window, textvariable = var1, bg = '#FAB60C')
label2.config(font=("Courier", 20))
var1.set('User Said:')
label2.pack()

label1 = Label(window, textvariable = var, bg = '#ADD8E6')
label1.config(font=("Courier", 20))
var.set('Welcome')
label1.pack()

frames = [PhotoImage(file='Assistant.gif',format = 'gif -index %i' %(i)) for i in range(100)]
window.title('Futuristic record assistance')

label = Label(window, width = 500, height = 500)
label.pack()
window.after(0, update, 0)
"""
btn0 = Button(text = 'WISH ME',width = 20, command = wishme, bg = '#5C85FB')
btn0.config(font=("Courier", 12))
btn0.pack()
"""
btn1 = Button(text = 'PLAY',width = 20,command = play, bg = '#5C85FB')
btn1.config(font=("Courier", 12))
btn1.pack()
btn2 = Button(text = 'EXIT',width = 20, command = window.destroy, bg = '#5C85FB')
btn2.config(font=("Courier", 12))
btn2.pack()


window.mainloop()